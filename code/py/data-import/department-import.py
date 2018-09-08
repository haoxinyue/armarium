from openpyxl import load_workbook
import psycopg2



wb = load_workbook(filename = 'F:\\Projects\\nick\\armarium\\doc\\测试数据\\太仓第一人民医院设备分析.xlsx', read_only=True)
sheet = wb['raw data']

department = {}

for row in range(2,sheet.max_row+1):
    cell_name = "{}{}".format("F", row)
    department[(sheet[cell_name].value)] = 1


conn = psycopg2.connect(host="47.100.198.255", dbname="armarium", user="armarium", password="armarium")
cur = conn.cursor()

for d in department.keys():
    sql = "INSERT INTO public.tb_department(	dept_name, dept_owner, dept_path, parent_id, hospital_id, remark, create_time, creater, modify_time, modifier) VALUES ('" + d + "', null, 1, 1000003, 1000003, null, '2018-09-06', 10001, '2018-09-06', 10001);"
    cur.execute(sql)
    conn.commit()


print("done")





