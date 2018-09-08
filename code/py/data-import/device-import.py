from openpyxl import load_workbook
import psycopg2

conn = psycopg2.connect(host="47.100.198.255", dbname="armarium", user="armarium", password="armarium")
cur = conn.cursor()

cur.execute("select * from tb_department where hospital_id = 1000003")
rows1 = cur.fetchall()

dict_department = {}

for row in rows1:
    dict_department[row[1]] = int(row[0])

dict_device_type = {}

cur.execute("select config_id,config_value from tsys_config where config_key = 'device_type'")
rows2 = cur.fetchall()
for row in rows2:
    dict_device_type[row[1]] = int(row[0])


wb = load_workbook(filename = 'F:\\Projects\\nick\\armarium\\doc\\测试数据\\太仓第一人民医院设备分析.xlsx', read_only=True)
sheet = wb['raw data']




insertSql = "INSERT INTO public.tb_device(device_code, device_name, hospital_id, dept_id, picture1, picture2, picture3, picture4, picture5, asset_no, device_model, device_desc, device_state, device_type, serial_number, usage_state, qr_code, manufacturer, producing_place, setup_date, accessory, accept_date, accept_remark, accept_file, use_date, device_owner, storage_date, purchase_amount, maintenance_end_date, warranty_begin_date, warranty_end_date, warranty_amount, warranty_content, " \
            "sales_supplier, sales_supplier_contact, sales_supplier_phone, sales_supplier_desc, after_sale_provider, after_sale_provider_engineer, after_sale_provider_phone, after_sale_provider_desc, contract_id, create_time, creater, modify_time, modifier, setup_case_id, need_inspection, inspection_interval, next_inspection_date, need_metering, metering_interval, next_metering_date, need_maintain, maintenance_interval, next_maintenance_date, inspection_owner)	" \
            "VALUES (null, %s, 1000003, %s, null, null, null, null, null, null, %s, null, null, %s, null, 1, null, %s, null, %s, null, null, null, null, %s, null, null, null, null, null, null, null, null, null, null, null, null, null, null, null, null, null, '2018-09-06', 10001, '2018-09-06', 10001, null, null, null, null, null, null, null, %s, %s, null, null);"
for row in range(2,sheet.max_row+1):

    id = sheet.cell(row=row,column=1).value

    device_name = sheet.cell(row=row,column=2).value
    device_model = sheet.cell(row=row,column=3).value

    dept = sheet.cell(row=row,column=6).value
    dept_id = dict_department[dept]

    device_type = sheet.cell(row=row,column=13).value
    device_type_id = dict_device_type[device_type]

    #serial_number = ?
    manufacturer = sheet.cell(row=row,column=4).value

    setup_date = str(sheet.cell(row=row,column=7).value).replace(".","-")
    use_date = str(sheet.cell(row=row, column=8).value).replace(".", "-")

    need_maintain = 1 if sheet.cell(row=row, column=17).value == 'Y' else 0

    maintenance_interval = None

    if need_maintain == 1:
        maintenance_interval = 360 / sheet.cell(row=row, column=18).value

    print(id, device_name, dept_id,device_model,device_type_id,manufacturer,setup_date,use_date,need_maintain,maintenance_interval)
    cur.execute(insertSql, (device_name, dept_id,device_model,device_type_id,manufacturer,setup_date,use_date,need_maintain,maintenance_interval))
conn.commit()
